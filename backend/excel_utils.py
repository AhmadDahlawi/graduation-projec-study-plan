import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re

# Excel template structure
TEMPLATE_COLUMNS = [
    'Plan Name',
    'Max Credits Per Semester',
    'Semester',
    'Course Name (Arabic)',
    'Course Name (English)',
    'Course Code',
    'Credit Hours',
    'Course Type',
    'Study Mode',
    'Lecture Hours',
    'Lab Hours',
    'Training Hours',
    'Department',
    'Prerequisite Codes'
]

def create_excel_template():
    """Create an Excel template for plan creation"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plan Template"
        
        # Add header row with styling
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, column_title in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = str(column_title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        column_widths = [20, 25, 12, 25, 25, 15, 15, 15, 15, 15, 15, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add sample data row
        sample_data = [
            "Sample Plan", "18", "1", "مقرر عينة", "Sample Course",
            "CS101", "3", "Required", "In-Person", "3", "0", "0", "Computer Science", ""
        ]
        
        for col_num, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
        
        # Convert to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        print(f"Error in create_excel_template: {e}")
        raise e

def validate_excel_file(file_content):
    """Validate Excel file format and structure"""
    try:
        df = pd.read_excel(file_content)
        missing_columns = [col for col in TEMPLATE_COLUMNS if col not in df.columns]
        if missing_columns:
            return False, f"Missing required columns: {', '.join(missing_columns)}"
        if df.empty:
            return False, "Excel file is empty."
        return True, df
    except Exception as e:
        return False, f"Failed to read Excel file: {str(e)}"

def validate_plan_data(df):
    """Validate plan data from Excel"""
    errors = []
    plan_names = df['Plan Name'].dropna().unique()
    if len(plan_names) == 0:
        errors.append("Plan Name is required")
    elif len(plan_names) > 1:
        errors.append("All rows must have the same Plan Name")
    
    for idx, row in df.iterrows():
        row_num = idx + 2
        if pd.isna(row['Semester']): errors.append(f"Row {row_num}: Semester is required")
        if pd.isna(row['Course Name (Arabic)']): errors.append(f"Row {row_num}: Course Name (Arabic) is required")
        if pd.isna(row['Course Name (English)']): errors.append(f"Row {row_num}: Course Name (English) is required")
        if pd.isna(row['Course Code']): errors.append(f"Row {row_num}: Course Code is required")
        if pd.isna(row['Credit Hours']): errors.append(f"Row {row_num}: Credit Hours is required")
    
    return errors

def export_plan_to_excel(plan_data, courses_data):
    """Export a plan to Excel format"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plan Data"
        
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num, column_title in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = str(column_title)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row_num = 2
        for course in courses_data:
            ws.cell(row=row_num, column=1).value = str(plan_data['name'])
            ws.cell(row=row_num, column=2).value = plan_data['max_credits_per_semester']
            ws.cell(row=row_num, column=3).value = course['semester']
            ws.cell(row=row_num, column=4).value = str(course['name_ar'])
            ws.cell(row=row_num, column=5).value = str(course['name_en'])
            ws.cell(row=row_num, column=6).value = str(course['code'])
            ws.cell(row=row_num, column=7).value = course['credits']
            ws.cell(row=row_num, column=8).value = str(course['type'])
            ws.cell(row=row_num, column=9).value = str(course['mode'])
            ws.cell(row=row_num, column=10).value = course.get('lecture_hours', 0)
            ws.cell(row=row_num, column=11).value = course.get('lab_hours', 0)
            ws.cell(row=row_num, column=12).value = course.get('training_hours', 0)
            ws.cell(row=row_num, column=13).value = str(course.get('department', ''))
            prereqs = course.get('prerequisite_codes', [])
            ws.cell(row=row_num, column=14).value = ', '.join(prereqs) if prereqs else ''
            
            for col in range(1, len(TEMPLATE_COLUMNS) + 1):
                ws.cell(row=row_num, column=col).border = border
            row_num += 1
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        print(f"Error in export_plan_to_excel: {e}")
        raise e
